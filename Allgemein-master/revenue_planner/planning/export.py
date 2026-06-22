"""Generate Excel output from planning results."""
from __future__ import annotations

import io
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from planning.engine import DayPlan

WEEKDAY_DE = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

# Colour palette
COL_HEADER  = "1F3864"   # dark blue
COL_SUB     = "2E75B6"   # mid blue
COL_FEIERT  = "FFE699"   # yellow
COL_FERIEN  = "C6EFCE"   # green
COL_SONDER  = "FCE4D6"   # orange
COL_CLOSED  = "D9D9D9"   # grey
COL_ALT     = "EBF3FB"   # light blue alt row
COL_WHITE   = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

EUR = '#,##0.00\\ "€"'


def _hfill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _hfont(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size)


def build_excel(results: list[DayPlan], gmbh_name: str, planjahr: int) -> bytes:
    """Return an Excel workbook as bytes."""
    df = pd.DataFrame([
        {
            "fil_nr": r.fil_nr,
            "datum": r.datum,
            "monat": r.datum.month,
            "wochentag": WEEKDAY_DE[r.wochentag],
            "tagestyp": r.tagestyp,
            "feiertag": r.feiertag_name,
            "ferien": r.ferien_art,
            "ist_vj": r.ist_vj,
            "monatsumsatz_ist_hoch": r.monat_hoch,
            "monatsumsatz_plan": r.monat_plan,
            "tagesumsatz_plan": r.budget,
            "gesamt_plan": r.budget,
            "normalisierung": r.normalisierung,
        }
        for r in results
    ])

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _sheet_jahresuebersicht(wb, df, planjahr, gmbh_name)
    _sheet_monatsuebersicht(wb, df, planjahr)
    _sheet_tagesdetail(wb, df, planjahr)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sheet 1: Jahresübersicht ───────────────────────────────────────────────

def _sheet_jahresuebersicht(wb: Workbook, df: pd.DataFrame, planjahr: int, gmbh: str):
    ws = wb.create_sheet("Jahresübersicht")

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Jahresplanung {planjahr} – {gmbh}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = _hfill(COL_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Filiale", "IST VJ (€)", "Budget (€)", "Δ (€)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = _hfont(bold=True, color="FFFFFF")
        c.fill = _hfill(COL_SUB)
        c.alignment = Alignment(horizontal="center")
        c.border = border

    summary = df.groupby("fil_nr").agg(
        ist_vj=("ist_vj", "sum"),
        gesamt_plan=("gesamt_plan", "sum"),
    ).reset_index().sort_values("fil_nr")
    summary["delta"] = summary["gesamt_plan"] - summary["ist_vj"]

    for row_i, row in enumerate(summary.itertuples(), 3):
        fill = _hfill(COL_ALT) if row_i % 2 == 0 else _hfill(COL_WHITE)
        vals = [row.fil_nr, row.ist_vj, row.gesamt_plan, row.delta]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.fill = fill
            c.border = border
            if col > 1:
                c.number_format = EUR

    # Totals
    total_row = row_i + 1
    ws.cell(row=total_row, column=1, value="GESAMT").font = _hfont(bold=True)
    for col, col_name in enumerate(["ist_vj", "gesamt_plan", "delta"], 2):
        c = ws.cell(row=total_row, column=col, value=summary[col_name].sum())
        c.font = _hfont(bold=True)
        c.number_format = EUR
        c.fill = _hfill(COL_HEADER)
        c.font = Font(bold=True, color="FFFFFF")

    _auto_col_width(ws)


# ── Sheet 2: Monatsübersicht ───────────────────────────────────────────────

MONTH_DE = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun",
            "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]


def _sheet_monatsuebersicht(wb: Workbook, df: pd.DataFrame, planjahr: int):
    ws = wb.create_sheet("Monatsübersicht")

    ws.merge_cells("A1:N1")
    ws["A1"] = f"Monatsplanung {planjahr}"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = _hfill(COL_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    header_cols = ["Filiale"] + MONTH_DE + ["Gesamt"]
    for col, h in enumerate(header_cols, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = _hfont(bold=True, color="FFFFFF")
        c.fill = _hfill(COL_SUB)
        c.alignment = Alignment(horizontal="center")
        c.border = border

    monthly = df.groupby(["fil_nr", "monat"])["gesamt_plan"].sum().unstack(fill_value=0)

    for row_i, (fil_nr, row) in enumerate(monthly.iterrows(), 3):
        fill = _hfill(COL_ALT) if row_i % 2 == 0 else _hfill(COL_WHITE)
        ws.cell(row=row_i, column=1, value=fil_nr).fill = fill
        for m in range(1, 13):
            c = ws.cell(row=row_i, column=m + 1, value=row.get(m, 0.0))
            c.fill = fill
            c.number_format = EUR
            c.border = border
        total_c = ws.cell(row=row_i, column=14, value=row.sum())
        total_c.fill = fill
        total_c.number_format = EUR
        total_c.font = _hfont(bold=True)
        total_c.border = border

    _auto_col_width(ws)


# ── Sheet 3: Tagesdetail ──────────────────────────────────────────────────

TYPE_COLORS = {
    "feiertag":   COL_FEIERT,
    "ferien":     COL_FERIEN,
    "sondertag":  COL_SONDER,
    "geschlossen": COL_CLOSED,
}


def _sheet_tagesdetail(wb: Workbook, df: pd.DataFrame, planjahr: int):
    ws = wb.create_sheet("Tagesdetail")

    ws.merge_cells("A1:K1")
    ws["A1"] = f"Tagesgenaue Planung {planjahr}"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = _hfill(COL_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    headers = ["Filiale", "Datum", "Wochentag", "Typ", "Feiertag/Sondertag",
               "Ferien", "IST VJ (€)", "Budget (€)", "Norm.-Faktor"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = _hfont(bold=True, color="FFFFFF")
        c.fill = _hfill(COL_SUB)
        c.alignment = Alignment(horizontal="center")
        c.border = border

    df_sorted = df.sort_values(["fil_nr", "datum"])
    for row_i, row in enumerate(df_sorted.itertuples(), 3):
        typ = row.tagestyp
        base_fill = TYPE_COLORS.get(typ, COL_WHITE if row_i % 2 == 0 else COL_ALT)
        fill = _hfill(base_fill)

        vals = [
            row.fil_nr,
            row.datum if isinstance(row.datum, date) else pd.Timestamp(row.datum).date(),
            row.wochentag,
            row.tagestyp,
            row.feiertag,
            row.ferien,
            row.ist_vj,
            row.gesamt_plan,
            row.normalisierung,
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.fill = fill
            c.border = border
            if col == 2 and isinstance(v, date):
                c.number_format = "DD.MM.YYYY"
            elif col in (7, 8):
                c.number_format = EUR

    ws.auto_filter.ref = f"A2:I{ws.max_row}"
    ws.freeze_panes = "A3"
    _auto_col_width(ws)


def _auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)
